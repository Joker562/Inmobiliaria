"""Carga, normalizacion y manejo de imagenes para el catalogo de propiedades.

Este modulo es tolerante a archivos Excel "reales" y desordenados:
- Soporta multiples hojas por libro.
- Detecta automaticamente la fila de encabezados (no siempre es la fila 1).
- Mapea columnas con nombres variables a un esquema canonico fijo.
- Parsea precios/metros cuadrados con texto sucio ("No disponible",
  rangos "113 - 114.70 m2", "$3,080,000 MXN", etc).
- Extrae hipervinculos de Excel (una celda puede decir "Ver listado" pero
  apuntar a una URL real via hyperlink).
- Descarga y cachea en disco las imagenes referenciadas; si no hay imagen
  real disponible, resuelve a un placeholder generado localmente.
"""
from __future__ import annotations

import hashlib
import re
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
import requests
from PIL import Image, ImageDraw, ImageFont

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
CACHE_DIR = BASE_DIR / "assets" / "cache"
IMG_DIR = BASE_DIR / "assets" / "img"
PLACEHOLDER_PATH = IMG_DIR / "placeholder.png"

DEFAULT_SOURCE_CANDIDATES = [
    DATA_DIR / "propiedades.xlsx",
    DATA_DIR / "propiedades_ejemplo.xlsx",
]

REQUEST_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; CatalogoInmuebles/1.0)"}
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp")
IMAGE_HOST_HINTS = (
    "images.unsplash.com", "cloudinary.com", "imgur.com", "googleusercontent.com",
    "naventcdn.com", "fbcdn.net", "drive.google.com",
)

CANONICAL_COLUMNS = [
    "id", "nombre", "ciudad", "zona", "direccion", "precio", "precio_original",
    "moneda", "estado", "recamaras", "banos", "estacionamientos",
    "m2_construccion", "m2_terreno", "precio_m2", "mantenimiento",
    "descripcion", "imagen_url", "imagenes_urls", "enlace_externo", "ubicacion_filtro",
    "hoja_origen",
]
MAX_GALLERY_IMAGES = 6

# Orden de prioridad: el primer campo de la lista reclama primero la columna
# que matchee alguno de sus alias (busqueda por substring en el encabezado
# normalizado). Cada columna del Excel solo puede ser reclamada una vez.
ALIASES: list[tuple[str, list[str]]] = [
    ("precio_m2", ["precio/m2", "precio por m2", "preciom2", "precio m2"]),
    ("precio", ["precio", "monto", "valor"]),
    ("moneda", ["moneda", "divisa", "usd equiv"]),
    ("m2_construccion", ["m2 construccion", "m2 construida", "area construida", "superficie construida"]),
    ("m2_terreno", ["m2 terreno", "terreno", "lote", "superficie terreno"]),
    ("recamaras", ["recamara", "habitacion", "cuarto", "dormitorio", "bedroom"]),
    ("banos", ["bano", "bathroom"]),
    ("estacionamientos", ["estacionamiento", "cochera", "garage", "parking"]),
    ("mantenimiento", ["mantenimiento", "hoa", "cuota"]),
    ("_estado_col", ["estado", "categoria", "status"]),
    ("descripcion", ["descripcion", "detalle", "observacion", "nota"]),
    ("imagen_url", ["imagen", "foto", "photo", "picture"]),
    ("enlace_externo", ["enlace", "link", "url", "fuente", "listado", "anuncio"]),
    ("direccion", ["direccion", "domicilio", "address"]),
    ("zona", ["colonia", "zona", "barrio", "sector", "ubicacion", "desarrollo"]),
    ("ciudad", ["ciudad", "municipio", "city"]),
    ("nombre", ["nombre", "titulo", "modelo", "propiedad"]),
    ("id", ["id", "folio"]),
]
# Fallback especial para encabezados "m2" / "m²" sin calificar.
_BARE_M2_KEYWORDS = ["m2", "m²", "metros"]


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = "".join(
        c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c)
    )
    text = re.sub(r"\s+", " ", text)
    return text


def _map_columns(headers: list[str]) -> dict[int, str]:
    normalized = [_normalize_text(h) for h in headers]
    claimed: dict[int, str] = {}
    used_cols: set[int] = set()

    for field, keywords in ALIASES:
        for idx, header in enumerate(normalized):
            if idx in used_cols or not header:
                continue
            if any(kw in header for kw in keywords):
                claimed[idx] = field
                used_cols.add(idx)
                break

    # Encabezados "m2" / "m²" sueltos -> construccion primero, terreno despues.
    for idx, header in enumerate(normalized):
        if idx in used_cols or not header:
            continue
        if any(kw == header or kw in header for kw in _BARE_M2_KEYWORDS):
            target = "m2_construccion" if "m2_construccion" not in claimed.values() else "m2_terreno"
            claimed[idx] = target
            used_cols.add(idx)

    return claimed


def _detect_header_row(ws) -> int:
    """Devuelve el indice de fila (1-based, openpyxl) con mas coincidencias de alias."""
    best_row, best_score = 1, -1
    max_scan = min(12, ws.max_row)
    known_keywords = [kw for _, kws in ALIASES for kw in kws] + _BARE_M2_KEYWORDS
    for row in range(1, max_scan + 1):
        cells = [c.value for c in ws[row]]
        normalized = [_normalize_text(c) for c in cells]
        score = sum(1 for h in normalized if h and any(kw in h for kw in known_keywords))
        non_empty = sum(1 for h in normalized if h)
        if non_empty == 0:
            continue
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score >= 2 else 1


def _parse_number(raw: object) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().lower()
    if not text or "no disponible" in text or text in {"-", "n/a", "na"}:
        return None

    range_match = re.search(
        r"(\d[\d,\.]*)\s*[–—-]\s*\$?\s*(\d[\d,\.]*)", text
    )
    if range_match:
        try:
            a = float(range_match.group(1).replace(",", ""))
            b = float(range_match.group(2).replace(",", ""))
            return round((a + b) / 2, 2)
        except ValueError:
            pass

    match = re.search(r"\d[\d,]*\.?\d*", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def _is_probable_image_url(url: str) -> bool:
    if not url:
        return False
    low = url.lower().split("?")[0]
    if low.endswith(IMAGE_EXTENSIONS):
        return True
    return any(host in url.lower() for host in IMAGE_HOST_HINTS)


def _cell_url(cell) -> Optional[str]:
    link = cell.hyperlink.target if cell.hyperlink else None
    if link:
        return link
    value = cell.value
    if isinstance(value, str) and value.strip().startswith("http"):
        return value.strip()
    return None


def _extract_rows(ws, header_row: int, col_map: dict[int, str], hoja_origen: str) -> list[dict]:
    imagen_col_idx = next((idx for idx, f in col_map.items() if f == "imagen_url"), None)
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_cells = ws[r]
        if all(c.value in (None, "") for c in row_cells):
            continue
        record: dict = {"hoja_origen": hoja_origen}
        raw_estado_val = None
        for idx, field in col_map.items():
            if idx >= len(row_cells):
                continue
            cell = row_cells[idx]
            value = cell.value

            if field in ("imagen_url", "enlace_externo"):
                url = _cell_url(cell)
                if not url:
                    continue
                # Independiente de si la columna se llama "imagen" o "enlace",
                # solo se trata como foto si la URL realmente parece una imagen
                # (por extension o por venir de un CDN de fotos conocido). Esto
                # evita que un "Ver fotos" que en realidad enlaza a un articulo
                # (no una foto) se cuele como imagen y se pierda el enlace.
                if _is_probable_image_url(url):
                    record["imagen_url"] = url
                else:
                    record["enlace_externo"] = url
                continue

            if field == "_estado_col":
                raw_estado_val = value
                continue

            record[field] = value

        # Fotos adicionales: columnas sin encabezado propio, contiguas a la
        # derecha de la columna de imagen principal, con valores tipo URL
        # (patron comun al pegar varias fotos de un mismo anuncio).
        imagenes_extra: list[str] = []
        if imagen_col_idx is not None:
            idx = imagen_col_idx + 1
            while idx < len(row_cells) and idx not in col_map:
                url = _cell_url(row_cells[idx])
                if not url:
                    break
                imagenes_extra.append(url)
                idx += 1
        record["_imagenes_extra"] = imagenes_extra

        if raw_estado_val is not None:
            record["_estado_raw"] = raw_estado_val
        rows.append(record)
    return rows


def _resolve_estado(df: pd.DataFrame) -> pd.Series:
    valid_labels = {"desarrollo", "mercado abierto", "preventa", "usado", "nuevo"}
    if "_estado_raw" in df.columns:
        normalized = df["_estado_raw"].map(_normalize_text)
        if normalized.isin(valid_labels).mean() > 0.5:
            return normalized.map(
                lambda v: "Desarrollo" if "desarrollo" in v or "preventa" in v or "nuevo" in v else "Mercado Abierto"
            )
    return df["hoja_origen"].map(
        lambda hoja: "Desarrollo" if "desarrollo" in _normalize_text(hoja) else "Mercado Abierto"
    )


def load_properties(source=None) -> pd.DataFrame:
    """Lee un Excel (path, bytes o file-like de Streamlit) y devuelve un
    DataFrame normalizado con las columnas de CANONICAL_COLUMNS."""
    if source is None:
        source = next((p for p in DEFAULT_SOURCE_CANDIDATES if p.exists()), None)
        if source is None:
            return pd.DataFrame(columns=CANONICAL_COLUMNS)

    wb = openpyxl.load_workbook(source, data_only=True)
    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header_row = _detect_header_row(ws)
        headers = [c.value for c in ws[header_row]]
        col_map = _map_columns(headers)
        if not col_map:
            continue
        all_rows.extend(_extract_rows(ws, header_row, col_map, sheet_name))

    if not all_rows:
        return pd.DataFrame(columns=CANONICAL_COLUMNS)

    df = pd.DataFrame(all_rows)

    if "precio" in df.columns:
        df["precio_original"] = df["precio"].astype(object)
        df["precio"] = df["precio"].map(_parse_number)
    else:
        df["precio_original"] = None
        df["precio"] = None

    for numeric_field in ("recamaras", "banos", "estacionamientos", "m2_construccion", "m2_terreno", "precio_m2"):
        if numeric_field in df.columns:
            df[numeric_field] = df[numeric_field].map(_parse_number)
        else:
            df[numeric_field] = None

    df["estado"] = _resolve_estado(df)

    for text_field in ("nombre", "ciudad", "zona", "direccion", "descripcion", "moneda", "imagen_url", "enlace_externo", "mantenimiento"):
        if text_field not in df.columns:
            df[text_field] = None

    df["zona"] = df["zona"].fillna("").astype(str).str.strip()
    df["ciudad"] = df["ciudad"].fillna("").astype(str).str.strip()
    df["ubicacion_filtro"] = df.apply(
        lambda row: row["zona"] or row["ciudad"] or "Sin ubicacion especificada", axis=1
    )

    df["nombre"] = df.apply(
        lambda row: str(row["nombre"]).strip() if row.get("nombre") not in (None, "", "nan") else f"Propiedad {row.name + 1} - {row['ubicacion_filtro']}",
        axis=1,
    )

    if "id" not in df.columns or df["id"].isna().any():
        df["id"] = range(1, len(df) + 1)

    def _build_gallery(row) -> list[str]:
        urls: list[str] = []
        cover = row.get("imagen_url")
        if isinstance(cover, str) and cover.startswith("http"):
            urls.append(cover)
        extra = row.get("_imagenes_extra") or []
        for u in extra:
            if u not in urls:
                urls.append(u)
        return urls[:MAX_GALLERY_IMAGES]

    if "_imagenes_extra" not in df.columns:
        df["_imagenes_extra"] = [[] for _ in range(len(df))]
    df["imagenes_urls"] = df.apply(_build_gallery, axis=1)

    for field in CANONICAL_COLUMNS:
        if field not in df.columns:
            df[field] = None

    df = df[CANONICAL_COLUMNS].reset_index(drop=True)
    df["recamaras"] = df["recamaras"].fillna(0)
    df["banos"] = df["banos"].fillna(0)
    df["estacionamientos"] = df["estacionamientos"].fillna(0)
    return df


THUMB_SIZE = (900, 600)  # relacion 3:2, igual que el placeholder


def _crop_to_fill(img: Image.Image, size: tuple[int, int] = THUMB_SIZE) -> Image.Image:
    """Recorta al centro y escala para que todas las fotos (retrato, paisaje,
    capturas irregulares) terminen con el mismo tamano/proporcion, sin
    deformarlas, y mantener parejas las tarjetas y el PDF."""
    target_w, target_h = size
    src_w, src_h = img.size
    target_ratio = target_w / target_h
    src_ratio = src_w / src_h
    if src_ratio > target_ratio:
        new_w = max(1, round(src_h * target_ratio))
        left = (src_w - new_w) // 2
        img = img.crop((left, 0, left + new_w, src_h))
    else:
        new_h = max(1, round(src_w / target_ratio))
        top = (src_h - new_h) // 2
        img = img.crop((0, top, src_w, top + new_h))
    return img.resize(size, Image.LANCZOS)


def get_placeholder_image() -> Path:
    """Genera (una sola vez) y devuelve la ruta de una imagen placeholder."""
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    if PLACEHOLDER_PATH.exists():
        return PLACEHOLDER_PATH

    img = Image.new("RGB", (900, 600), color=(226, 232, 240))
    draw = ImageDraw.Draw(img)
    text = "Imagen no disponible"
    try:
        font = ImageFont.truetype("arial.ttf", 36)
    except Exception:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), text, font=font)
    w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.rectangle([40, 40, 860, 560], outline=(148, 163, 184), width=4)
    draw.line([40, 40, 860, 560], fill=(148, 163, 184), width=3)
    draw.line([40, 560, 860, 40], fill=(148, 163, 184), width=3)
    draw.rectangle([(450 - w / 2 - 20, 300 - h / 2 - 12), (450 + w / 2 + 20, 300 + h / 2 + 12)], fill=(226, 232, 240))
    draw.text((450 - w / 2, 300 - h / 2), text, fill=(71, 85, 105), font=font)
    img.save(PLACEHOLDER_PATH)
    return PLACEHOLDER_PATH


def get_property_image(url: Optional[str], timeout: int = 6) -> Path:
    """Descarga (con cache en disco) la imagen de `url`. Si falla o no hay
    URL, devuelve la ruta del placeholder."""
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return get_placeholder_image()

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    key = hashlib.sha256(url.encode("utf-8")).hexdigest()[:24]
    ext = Path(url.split("?")[0]).suffix.lower()
    if ext not in IMAGE_EXTENSIONS:
        ext = ".jpg"
    cached_path = CACHE_DIR / f"{key}{ext}"

    if cached_path.exists() and cached_path.stat().st_size > 0:
        return cached_path

    try:
        resp = requests.get(url, headers=REQUEST_HEADERS, timeout=timeout)
        resp.raise_for_status()
        img = Image.open(pd.io.common.BytesIO(resp.content))
        img = _crop_to_fill(img.convert("RGB"))
        img.save(cached_path, format="JPEG", quality=85)
        return cached_path
    except Exception:
        return get_placeholder_image()


def format_currency(value: Optional[float], moneda: Optional[str] = None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "Precio no disponible"
    symbol = "USD" if moneda and "usd" in str(moneda).lower() else "MXN"
    return f"${value:,.0f} {symbol}"
